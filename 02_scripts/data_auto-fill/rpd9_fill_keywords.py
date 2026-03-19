"""
Script name: rpd9_fill_keywords.py

Description:
This script automatically generates keywords for each entry in the RPD dataset using the KeyBERT AI model. It combines text from multiple source columns and populates the "Keywords" column with the top AI-generated keywords. Rows where keywords could not be generated are highlighted in yellow, and a log file is created to track successes, failures, and skipped rows.

Keywords are extracted from the following columns:
- "Programme name"
- "Payor country or economy"
- "Recipient country or economy"
- "Context or reason for reparations"
- "Main harm"
- "Source extracts"

Author: Geraldine Ramilo  
Project: REPAIR  

Notes:
- Developed with AI-assisted support (ChatGPT)  
- Reviewed and tested by the author  
- Preserves existing Excel formatting  
- Creates a new output file and does not overwrite the original database  
- Only fills empty cells in the "Keywords" column (does not overwrite existing keywords)  
- Combines text from the above columns into a single string for keyword extraction  
- Highlights rows in yellow where no keywords could be generated  
- Requires an active internet connection to use the KeyBERT model  
- Currently processes the first 500 rows (adjustable in the script)  
"""

import pandas as pd
from keybert import KeyBERT
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Config ===
input_file = "RPD (V1).xlsx"
output_file = "RPD (V1) - updated.xlsx"
log_file = "keyword_log.csv"

# === Load Excel data ===
df = pd.read_excel(input_file, engine='openpyxl')

# === Safe row bounds ===
start_row = 0
max_rows = df.shape[0]
end_row = min(500, max_rows)  # adjust if needed

# === Load KeyBERT model ===
kw_model = KeyBERT(model='all-MiniLM-L6-v2')

# === Source columns by name ===
source_cols = [
    "Programme name",
    "Payor country or economy",
    "Recipient country or economy",
    "Context or reason for reparations",
    "Main harm",
    "Source extracts"
]

# === Output column ===
col_keywords = "Keywords"

# === Keyword extraction function ===
def extract_keywords(text, top_n=10):
    if not isinstance(text, str) or text.strip() == '':
        return ''
    keywords = kw_model.extract_keywords(text, stop_words='english', top_n=top_n)
    return ', '.join([kw[0] for kw in keywords])

# === Track logs and failures ===
log = []
rows_updated = 0

for idx in range(start_row, end_row):
    row_number = idx + 2  # Excel row numbers

    # Only process if Keywords cell is empty
    if pd.isna(df.at[idx, col_keywords]) or df.at[idx, col_keywords] == '':
        combined_text = ' '.join([
            str(df.at[idx, col]) if pd.notna(df.at[idx, col]) else ''
            for col in source_cols
        ]).strip()

        if combined_text:
            keywords = extract_keywords(combined_text)
            df.at[idx, col_keywords] = keywords
            log.append({'Row': row_number, 'Success': True, 'Keywords': keywords})
            rows_updated += 1
        else:
            df.at[idx, col_keywords] = ''
            log.append({'Row': row_number, 'Success': False, 'Keywords': ''})
    else:
        log.append({'Row': row_number, 'Success': 'Skipped', 'Keywords': df.at[idx, col_keywords]})

# === Save updated Excel ===
df.to_excel(output_file, index=False)

# === Highlight failed keyword rows (yellow fill in Keywords column) ===
wb = load_workbook(output_file)
ws = wb.active
keywords_col_index = df.columns.get_loc(col_keywords) + 1  # 1-indexed for openpyxl

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for entry in log:
    if entry['Success'] is False:
        ws.cell(row=entry['Row'], column=keywords_col_index).fill = yellow_fill

wb.save(output_file)

# === Save log to CSV ===
log_df = pd.DataFrame(log)
log_df.to_csv(log_file, index=False)

print(f"✅ Done! Keywords added to '{output_file}', failed rows highlighted, and log saved to '{log_file}'.")
print(f"ℹ️ Rows updated with new keywords: {rows_updated}")
