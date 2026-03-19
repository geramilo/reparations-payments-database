"""
Script name: rpd8_archive_waybackmachine.py

Description:
This script retrieves the latest available archived versions of source URLs using the Wayback Machine API and records them in the RPD. It processes the "Source URL(s)" column and fills the corresponding "Archived URL(s) via Wayback Machine" column with formatted archive links and dates.

Author: Geraldine Ramilo
Project: REPAIR  

Notes:
- Developed with AI-assisted support (ChatGPT)  
- Reviewed and tested by the author  
- Preserves existing Excel formatting  
- Creates a new output file and does not overwrite the original database  
- Only fills archive cells that are currently empty (existing data is preserved)  
- Supports multiple URLs per cell (comma-separated)  
- Requires an active internet connection to query the Wayback Machine API  
- Please ensure that the input file is located in the specified directory before execution  
"""


import os
import requests
import openpyxl
from urllib.parse import urlparse
from datetime import datetime

# File paths
desktop = os.path.expanduser("~/Desktop")
input_file = os.path.join(desktop, "database", "RPD (V1).xlsx")
output_file = os.path.join(desktop, "database", "RPD (V1) - updated.xlsx")

# Load workbook
wb = openpyxl.load_workbook(input_file)
ws = wb["DATABASE"]

# Find column indices
header = {cell.value: cell.column for cell in ws[1]}
source_col = header.get("Source URL(s)")
archive_col = header.get("Archived URL(s) via Wayback Machine")

if not source_col or not archive_col:
    raise ValueError("Required columns not found in the sheet.")

# Wayback Machine API
WAYBACK_API = "http://archive.org/wayback/available"

def get_latest_snapshot(url):
    try:
        r = requests.get(WAYBACK_API, params={"url": url}, timeout=10)
        data = r.json()
        snapshots = data.get("archived_snapshots", {})
        if "closest" in snapshots:
            snap = snapshots["closest"]
            archived_url = snap["url"]
            date_str = snap["timestamp"][:8]  # YYYYMMDD
            dt = datetime.strptime(date_str, "%Y%m%d")
            formatted_date = dt.strftime("%d %B %Y")
            domain = urlparse(url).netloc
            return f"{domain}: {archived_url} [{formatted_date}]"
        else:
            return "Not available"
    except Exception:
        return "Not available"

# Iterate rows with progress logging
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    source_cell = row[source_col - 1]
    archive_cell = row[archive_col - 1]

    if source_cell.value and not archive_cell.value:
        print(f"Processing row {i} ...")
        urls = [u.strip() for u in source_cell.value.split(",") if u.strip()]
        results = []
        for url in urls:
            print(f"   Checking {url}")
            results.append(get_latest_snapshot(url))
        archive_cell.value = "\n\n".join(results)

# Save new file
wb.save(output_file)
print(f"\n✅ Finished! Saved to {output_file}")
