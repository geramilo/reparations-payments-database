"""
Script name: rpd1_assign_id.py

Description:
This script assigns a unique ID to each row of the RPD based on:
- ISO code of the paying country
- ISO code of the recipient country
- Year of the reparations agreement
- A sequential counter for duplicate cases

Example ID format:
DEU-NAM-2004-001

Author: Geraldine Ramilo
Project: REPAIR

Notes:
- Developed with AI-assisted support (ChatGPT)
- Reviewed and tested by the author
- Preserves existing Excel formatting
- Creates a new output file and does not overwrite the original database
"""

from pathlib import Path
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# ===== CONFIGURATION =====
# Update these paths to match your folder structure
BASE_DIR = Path.home() / # Place the input file in the same folder as this script
INPUT_FILE = BASE_DIR / "RPD (V1).xlsx"
OUTPUT_FILE = BASE_DIR / "RPD (V1) updated.xlsx"

SHEET_NAME = "DATABASE"

REQUIRED_COLUMNS = {
    "ID",
    "Payor country or economy (ISO)",
    "Recipient country or economy (ISO)",
    "Year of reparations agreement",
}


# ===== LOAD WORKBOOK =====
wb = load_workbook(INPUT_FILE)
ws = wb[SHEET_NAME]

# ===== IDENTIFY REQUIRED COLUMNS =====
headers = {cell.value: cell.column for cell in ws[1]}

id_col = headers.get("ID")
payor_col = headers.get("Payor country or economy (ISO)")
recipient_col = headers.get("Recipient country or economy (ISO)")
year_col = headers.get("Year of reparations agreement")

if not all([id_col, payor_col, recipient_col, year_col]):
    raise ValueError("One or more required columns are missing!")

# ===== ASSIGN UNIQUE IDS =====
counter = {} # track occurrences for sequential numbering

for row in range(2, ws.max_row + 1):  # Skip header row
    payor = ws.cell(row=row, column=payor_col).value
    recipient = ws.cell(row=row, column=recipient_col).value
    year = ws.cell(row=row, column=year_col).value

    # Handle missing or "Unknown" year
    if not year or str(year).strip().lower() == "unknown":
        year_str = "XXXX"
    else:
        year_str = str(year)

    # Build base key
    base_key = f"{payor}-{recipient}-{year_str}"

    # Count occurrence
    counter[base_key] = counter.get(base_key, 0) + 1
    seq = f"{counter[base_key]:03d}"  # Format as 001, 002, ...

    # Final ID
    final_id = f"{base_key}-{seq}"

    # Write to first column ("ID")
    ws.cell(row=row, column=id_col, value=final_id)

    # Preserve bold if header was bold
    if row == 1:
        ws.cell(row=row, column=id_col).font = Font(bold=True)

# ===== SAVE OUTPUT =====
wb.save(OUTPUT_FILE)

print(f"✅ IDs populated and saved as: {OUTPUT_FILE}")
