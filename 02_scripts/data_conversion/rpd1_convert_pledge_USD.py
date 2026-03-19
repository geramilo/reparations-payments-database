"""
==================================================================
RPD Currency Conversion Script

Script name:rpd1_convert_pledge_USD.py

Author: Geraldine Ramilo  
Project: REPAIR  

Notes:
- Developed with AI-assisted support (ChatGPT)  
- Reviewed and tested by the author 
==================================================================

Purpose of the script:
--------
This script converts the 'Pledged (Total)' amounts in the 
RPD from local currencies to USD using historical 
exchange rates.

Input Files:
------------
1. RPD (V1).xlsx
   - Sheet: DATABASE
   - Columns required:
     * 'Pledged (Total)'  → amount to convert
     * 'Currency'         → ISO currency code (e.g., USD, EUR)
     * 'Year of reparations agreement' → year for exchange rate

2. Exhange Rates Data (Processed).xlsx
   - Sheet: LCU per USD, 1945-2024
   - Column D: 'LCU (ISO)' → ISO currency codes
   - Columns E→CF: annual exchange rates from 1945 to 2024

Output:
-------
- RPD (V1) - CONVERTED.xlsx
- Adds USD conversions in 'Pledged (Total USD)'
- Adds a 'Conversion Notes' sheet with logs, errors, and metadata

Key Features:
-------------
- Skips rows already filled in USD
- Marks missing/invalid data or missing exchange rates in YELLOW
- Logs all updates and issues in 'Conversion Notes'

Things to Remember:
-------------------
- Make sure the script and both input Excel files are in the same directory where you run the terminal.
- Currency codes in the RPD must be ISO 3-letter codes
- Only works for years covered in the exchange rate sheet (1945-2024)
- Always run a backup of your original REPAIR Database before converting
- The script rounds USD values to 2 decimal places
- Missing or invalid data will not be converted, but logged in yellow
"""


import os
import re
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

# ---------- CONFIG ----------
# Use current working directory (where you run the script)
folder = os.getcwd()  # relative to terminal directory

repair_file = os.path.join(folder, "RPD (V1).xlsx")
money_file = os.path.join(folder, "Exhange Rates Data (Processed).xlsx")
repair_output = os.path.join(folder, "RPD (V1) - CONVERTED.xlsx")

repair_sheet_name = "DATABASE"
money_sheet_name = "LCU per USD, 1945-2024"

HDR_AMOUNT = "Pledged (Total)"
HDR_CURRENCY = "Currency"
HDR_YEAR = "Year of reparations agreement"
HDR_USD = "Pledged (Total USD)"

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ---------- CURRENCY MAP ----------
currency_map = {
    "US Dollar": "USD",
    "Euro": "EUR",
    "Japanese Yen": "JPY",
    "British Pound": "GBP",
    # Add more mappings as needed
}

# ---------- HELPERS ----------
def parse_year(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    m = re.search(r"\d{4}", s)
    return int(m.group(0)) if m else None

def parse_amount(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "").replace(" ", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except:
        return None

def normalize_currency(val):
    if val is None:
        return None
    s = str(val).strip()
    # Exact 3-letter ISO match
    if re.fullmatch(r"[A-Z]{3}", s.upper()):
        return s.upper()
    # Check map
    return currency_map.get(s, None)

# ---------- LOAD WORKBOOKS ----------
repair_wb = openpyxl.load_workbook(repair_file)
repair_ws = repair_wb[repair_sheet_name]

money_wb = openpyxl.load_workbook(money_file, data_only=True)
money_ws = money_wb[money_sheet_name]

# ---------- HEADER INDICES ----------
headers = {cell.value: idx for idx, cell in enumerate(next(repair_ws.iter_rows(min_row=1, max_row=1)), start=1)}

col_amount = headers.get(HDR_AMOUNT)
col_currency = headers.get(HDR_CURRENCY)
col_year = headers.get(HDR_YEAR)
col_usd = headers.get(HDR_USD)

if not all([col_amount, col_currency, col_year, col_usd]):
    raise ValueError("❌ One or more required headers not found in Repair Database.")

# ---------- BUILD EXCHANGE RATES ----------
# Read year headers E→CF (columns 5→84)
money_header = [cell.value for cell in money_ws[1]]
year_cols = {}
for idx, val in enumerate(money_header[4:], start=4):
    if val is None:
        continue
    try:
        year_cols[int(str(val).strip())] = idx
    except:
        continue

exchange_rates = {}
for row in money_ws.iter_rows(min_row=2, values_only=True):
    raw_currency = row[3]  # column D
    currency = normalize_currency(raw_currency)
    if not currency:
        continue
    exchange_rates.setdefault(currency, {})
    for year, idx in year_cols.items():
        if idx < len(row):
            rate = row[idx]
            if isinstance(rate, (int, float)) and rate != 0:
                exchange_rates[currency][year] = float(rate)

# ---------- PROCESS RPD DATA ----------
log = []
updates = 0

for r in range(2, repair_ws.max_row + 1):
    amount_cell = repair_ws.cell(r, col_amount)
    currency_cell = repair_ws.cell(r, col_currency)
    year_cell = repair_ws.cell(r, col_year)
    usd_cell = repair_ws.cell(r, col_usd)

    if usd_cell.value is not None:
        continue  # Skip already filled

    amount = parse_amount(amount_cell.value)
    currency = normalize_currency(currency_cell.value)
    year = parse_year(year_cell.value)

    if not all([amount, currency, year]):
        usd_cell.fill = yellow_fill
        log.append(f"Row {r}: Missing/invalid data -> amount:{amount_cell.value}, currency:{currency_cell.value}, year:{year_cell.value}")
        continue

    rate = exchange_rates.get(currency, {}).get(year)
    if not rate:
        usd_cell.fill = yellow_fill
        log.append(f"Row {r}: No exchange rate for {currency} in {year}")
        continue

    try:
        usd_value = round(amount / rate, 2)
        usd_cell.value = usd_value
        updates += 1
        log.append(f"Row {r}: {amount} {currency} ({year}) -> {usd_value} USD")
    except Exception as e:
        usd_cell.fill = yellow_fill
        log.append(f"Row {r}: Conversion failed ({e})")

# ---------- ADD NOTES SHEET ----------
note_ws = repair_wb.create_sheet("Conversion Notes")
note_ws["A1"] = "Conversion completed on " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
note_ws["A2"] = f"Source file: {os.path.basename(repair_file)}"
note_ws["A3"] = f"Exchange rates: {os.path.basename(money_file)}"
note_ws["A4"] = f"Output column: {HDR_USD}"
note_ws["A5"] = f"Rows updated: {updates}"
note_ws["A7"] = "Detailed log:"
for i, entry in enumerate(log, start=8):
    note_ws[f"A{i}"] = entry

# ---------- SAVE ----------
repair_wb.save(repair_output)
print(f"✅ Conversion complete. Saved to {repair_output}")
print(f"Rows updated: {updates}")
print(f"Check 'Conversion Notes' sheet for full log.")
