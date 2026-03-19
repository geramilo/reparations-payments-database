"""
==================================================================
RPD Currency Conversion Script

Script name: rpd4_convert_paid_USD2024.py

Author: Geraldine Ramilo  
Project: REPAIR  

Notes:
- Developed with AI-assisted support (ChatGPT)  
- Reviewed and tested by the author
==================================================================

Purpose of the script:
--------
This script converts the 'Paid (Total USD)' amounts in the 
RPD database to their equivalent value in 2024 USD using the 
Consumer Price Index (CPI). It ensures all amounts are adjusted 
for inflation and highlights missing or problematic data.

Input Files:
------------
1. RPD (V1).xlsx
   - Sheet: DATABASE
   - Columns required:
     * 'Paid (Total USD)' → amount to convert
     * 'Payments conversion year' → year used for CPI adjustment
     * 'Paid (Total USD 2024)' → column to populate with converted values

2. CPI USD Data.xlsx
   - Sheet: CPI Data
   - Columns: Year → CPI value
   - Used to calculate equivalent amounts in 2024 USD

Output:
-------
- RPD (V1) - UPDATED.xlsx
- Populates 'Paid (Total USD 2024)' with converted amounts
- Adds a 'Conversion Notes' sheet that includes:
  * Conversion timestamp
  * Source files
  * Number of rows updated
  * Detailed log of all conversions, errors, and missing data

Key Features:
-------------
- Skips rows already filled in the target column
- Marks missing/invalid data or missing CPI values in YELLOW
- Logs all updates and issues in 'Conversion Notes'
- Cleans and parses numeric strings with commas (e.g., "1,000" → 1000)
- Ensures only empty cells are converted to avoid overwriting existing data

Things to Remember:
-------------------
- Highlights rows with missing or invalid amounts or missing CPI data in **yellow**
- Logs all updates, errors, and skipped rows in the 'Conversion Notes' sheet
- Uses current working directory for all input/output files, making it portable across systems
"""


import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os

# File paths
cwd = os.getcwd()  # current working directory

database_file = os.path.join(cwd, "RPD (V1).xlsx")
cpi_file = os.path.join(cwd, "CPI USD Data.xlsx")
output_file = os.path.join(cwd, "RPD (V1) - UPDATED.xlsx")

# Load workbooks
db_wb = openpyxl.load_workbook(database_file)
db_ws = db_wb["DATABASE"]

cpi_wb = openpyxl.load_workbook(cpi_file)
cpi_ws = cpi_wb["CPI Data"]

# Build CPI dictionary: year -> CPI
cpi_dict = {}
for row in cpi_ws.iter_rows(min_row=2, values_only=True):
    year, cpi = row
    cpi_dict[year] = cpi

# Create a yellow fill for missing values
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Find the column numbers for the relevant data
header_row = 1
col_amount = None
col_year = None
col_adjusted = None

for col in range(1, db_ws.max_column + 1):
    header = db_ws.cell(row=header_row, column=col).value
    if header == "Paid (Total USD)":
        col_amount = col
    elif header == "Payments conversion year":
        col_year = col
    elif header == "Paid (Total USD 2024)":
        col_adjusted = col

if col_amount is None or col_year is None or col_adjusted is None:
    raise ValueError("Could not find the required columns in DATABASE sheet.")

# Create or clear "Conversion Notes" sheet
if "Conversion Notes" in db_wb.sheetnames:
    db_wb.remove(db_wb["Conversion Notes"])
notes_ws = db_wb.create_sheet("Conversion Notes")

# Track log and updates
log = []
updates = 0

# Perform conversion only for EMPTY cells in "Paid (Total USD 2024)"
for row in range(2, db_ws.max_row + 1):
    adjusted_cell = db_ws.cell(row=row, column=col_adjusted)

    # Skip non-empty cells
    if adjusted_cell.value is not None:
        continue

    amount = db_ws.cell(row=row, column=col_amount).value
    year = db_ws.cell(row=row, column=col_year).value

    if amount is None or year not in cpi_dict:
        adjusted_cell.fill = yellow_fill
        log.append(f"Row {row}: Could not convert (missing amount or CPI for year {year})")
        continue

    try:
        # Clean and convert to float
        if isinstance(amount, str):
            amount = amount.replace(",", "").strip()
        original_amount = amount
        amount = float(amount)
    except Exception:
        adjusted_cell.fill = yellow_fill
        log.append(f"Row {row}: Could not convert value '{amount}' (highlighted in yellow)")
        continue

    cpi_year = cpi_dict[year]
    cpi_2024 = cpi_dict.get(2024)

    if cpi_2024 is None:
        raise ValueError("CPI for 2024 not found in CPI USD Data.xlsx")

    adjusted_value = amount * (cpi_2024 / cpi_year)
    adjusted_cell.value = round(adjusted_value, 2)

    log.append(
        f"Row {row}: {original_amount} (year {year}) → {round(adjusted_value, 2)} (2024 USD)"
    )
    updates += 1

# ---------- ADD NOTES SHEET ----------
notes_ws["A1"] = "Conversion completed on " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
notes_ws["A2"] = f"Source file: {os.path.basename(database_file)}"
notes_ws["A3"] = f"CPI data: {os.path.basename(cpi_file)}"
notes_ws["A4"] = f"Output column: Paid (Total USD 2024)"
notes_ws["A5"] = f"Rows updated: {updates}"
notes_ws["A7"] = "Detailed log:"

for i, entry in enumerate(log, start=8):
    notes_ws[f"A{i}"] = entry

# ---------- SAVE ----------
db_wb.save(output_file)
print(f"✅ Conversion complete. Saved to {output_file}")
print(f"Rows updated: {updates}")
print(f"Check 'Conversion Notes' sheet for full log.")
